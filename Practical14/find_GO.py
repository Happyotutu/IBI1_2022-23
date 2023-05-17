import xml.dom.minidom
from openpyxl import Workbook

DOM = xml.dom.minidom.parse("go_obo.xml")
term_elements = DOM.getElementsByTagName('term')
workbook = Workbook()
sheet = workbook.active
# Write headers to the Excel sheet
sheet['A1'] = 'GO ID'
sheet['B1'] = 'Term Name'
sheet['C1'] = 'Definition'
sheet['D1'] = 'Number of Child Nodes'
row = 2 # output from the row 2 because row1

def count_child_nodes(child_id):
    child_node_count = 0
    for term_element in term_elements:
        id_node_child = term_element.getElementsByTagName('id')[0]
        is_a_nodes = term_element.getElementsByTagName('is_a')
        for is_a_node in is_a_nodes:
            is_a_value = is_a_node.firstChild.nodeValue
            if child_id == is_a_value:
                child_node_count += count_child_nodes(id_node_child.firstChild.nodeValue) + 1            
    return child_node_count

# Iterate over the <term> elements
for term_element in term_elements:
    defstr_element = term_element.getElementsByTagName('defstr')[0]
    id_element = term_element.getElementsByTagName('id')[0]
    name_element = term_element.getElementsByTagName('name')[0]
    num_child_nodes = 0
    # Check if the text in <defstr> contains 'autophagosome'
    if 'autophagosome' in defstr_element.firstChild.data:
        num_child_nodes = count_child_nodes(id_element.firstChild.data)

        # Write data to the Excel sheet
        sheet.cell(row, column=1).value = id_element.firstChild.data
        sheet.cell(row, column=2).value = name_element.firstChild.data
        sheet.cell(row, column=3).value = defstr_element.firstChild.data
        sheet.cell(row, column=4).value = num_child_nodes
        row += 1
# Save the Excel workbook
workbook.save("autophagosome.xlsx")
